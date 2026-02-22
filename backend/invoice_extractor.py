# Invoice Extractor Module - Hybrid 4-Stage Pipeline
# Uses the new universal_extractor with template-specific parsing

from fastapi import APIRouter, UploadFile, File, HTTPException, BackgroundTasks
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone
import asyncio
import aiofiles
import json
import re
import csv

# Import new 4-stage pipeline
from universal_extractor import (
    extract_document_content,
    detect_template,
    universal_extract,
    universal_extract_async,
    normalize_amount,
    normalize_date,
    detect_platform,
    DocumentContent
)

# Excel processing
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

logger = logging.getLogger(__name__)

# File storage directories
ROOT_DIR = Path(__file__).parent
UPLOAD_DIR = ROOT_DIR / "uploads"
EXPORT_DIR = ROOT_DIR / "exports"
UPLOAD_DIR.mkdir(exist_ok=True)
EXPORT_DIR.mkdir(exist_ok=True)

# Max file size (25MB)
MAX_FILE_SIZE = 25 * 1024 * 1024

# Create router
invoice_router = APIRouter(prefix="/api/invoice")

# ============== PYDANTIC MODELS ==============

class LineItem(BaseModel):
    category_code_or_hsn: Optional[str] = None
    service_description: Optional[str] = None
    fee_amount: Optional[float] = None
    cgst_amount: Optional[float] = None
    sgst_amount: Optional[float] = None
    igst_amount: Optional[float] = None
    total_tax_amount: Optional[float] = None
    total_amount: Optional[float] = None
    tax_rate_percent: Optional[float] = None

class InvoiceData(BaseModel):
    source_platform: str = "Unknown"
    document_type: str = "Invoice"
    invoice_number: Optional[str] = None
    invoice_date: Optional[str] = None
    service_provider_name: Optional[str] = None
    service_provider_gstin: Optional[str] = None
    service_receiver_name: Optional[str] = None
    service_receiver_gstin: Optional[str] = None
    place_of_supply_state_code: Optional[str] = None
    currency: str = "INR"
    subtotal_fee_amount: Optional[float] = None
    cgst_amount: Optional[float] = None
    sgst_amount: Optional[float] = None
    igst_amount: Optional[float] = None
    total_tax_amount: Optional[float] = None
    total_invoice_amount: Optional[float] = None
    line_items: List[LineItem] = []

class UploadResponse(BaseModel):
    job_id: str
    files: List[Dict[str, str]]

class JobStatusResponse(BaseModel):
    job_id: str
    status: str
    total_files: int
    processed_files: int
    failed_files: int
    files: List[Dict[str, Any]]

# Database reference (will be set from main server)
db = None

def set_database(database):
    global db
    db = database

# ============== PDF TEXT EXTRACTION ==============

def extract_text_from_pdf(file_path: str) -> tuple:
    """
    Extract text from PDF using the 4-stage pipeline.
    Returns (full_text, pages_text, document_info)
    """
    doc_content = extract_document_content(file_path)
    
    document_info = {
        'extraction_method': doc_content.extraction_method,
        'is_scanned': doc_content.is_scanned,
        'confidence': doc_content.confidence,
        'page_count': doc_content.page_count,
        'tables_found': doc_content.tables_found,
        'anchors_found': doc_content.anchors_found
    }
    
    return doc_content.full_text, doc_content.pages_text, document_info


# ============== FILE PROCESSING ==============

async def process_single_file(job_id: str, file_info: Dict, use_llm: bool = True, api_key: str = None) -> None:
    """Process a single PDF file using the hybrid pipeline."""
    if db is None:
        logger.error("Database not initialized")
        return
        
    file_id = file_info['id']
    filename = file_info['filename']
    original_filename = file_info['original_filename']
    file_path = UPLOAD_DIR / filename
    
    try:
        await db.invoice_job_files.update_one(
            {"id": file_id, "job_id": job_id},
            {"$set": {"status": "processing", "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
        
        # Stage A: Extract document content
        text, pages_text, doc_info = extract_text_from_pdf(str(file_path))
        
        if not text or len(text.strip()) < 50:
            raise ValueError("Could not extract sufficient text from PDF")
        
        # Stage B, C, D: Extract using hybrid pipeline
        if use_llm and api_key:
            extraction_data = await universal_extract_async(
                text, original_filename, pages_text, 
                use_llm=use_llm, api_key=api_key
            )
        else:
            extraction_data = universal_extract(text, original_filename, pages_text)
        
        # Add document info
        extraction_data['_document_info'] = doc_info
        extraction_data['_source_file'] = original_filename
        
        await db.invoice_job_files.update_one(
            {"id": file_id, "job_id": job_id},
            {
                "$set": {
                    "status": "done",
                    "extraction_data": extraction_data,
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
            }
        )
        
        await db.invoice_jobs.update_one(
            {"job_id": job_id},
            {"$inc": {"processed_files": 1}}
        )
        
        template_id = extraction_data.get('_template_id', 'UNKNOWN')
        validation = extraction_data.get('_validation', {})
        logger.info(f"Processed {original_filename}: template={template_id}, status={validation.get('status', 'ok')}")
        
    except Exception as e:
        error_msg = str(e)
        logger.error(f"Failed to process {original_filename}: {error_msg}")
        
        await db.invoice_job_files.update_one(
            {"id": file_id, "job_id": job_id},
            {
                "$set": {
                    "status": "failed",
                    "error_message": error_msg,
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
            }
        )
        
        await db.invoice_jobs.update_one(
            {"job_id": job_id},
            {"$inc": {"failed_files": 1, "processed_files": 1}}
        )


async def run_extraction_job(job_id: str, use_llm: bool = True, api_key: str = None) -> None:
    """Run the complete extraction job."""
    if db is None:
        logger.error("Database not initialized")
        return
        
    try:
        files = await db.invoice_job_files.find({"job_id": job_id}, {"_id": 0}).to_list(1000)
        
        await db.invoice_jobs.update_one(
            {"job_id": job_id},
            {"$set": {"status": "processing"}}
        )
        
        # Process files sequentially to avoid rate limits
        for file_info in files:
            await process_single_file(job_id, file_info, use_llm, api_key)
        
        await db.invoice_jobs.update_one(
            {"job_id": job_id},
            {"$set": {"status": "completed"}}
        )
        
        logger.info(f"Invoice Job {job_id} completed")
        
    except Exception as e:
        logger.error(f"Invoice Job {job_id} failed: {e}")
        await db.invoice_jobs.update_one(
            {"job_id": job_id},
            {"$set": {"status": "failed"}}
        )


# ============== EXPORT HELPERS ==============

def generate_csv(job_id: str, files_data: List[Dict]) -> str:
    """Generate CSV file with line items (long format)."""
    csv_path = EXPORT_DIR / f"{job_id}_invoice_line_items.csv"
    
    headers = [
        "Invoice Number", "Invoice Date", "Document Type", "Platform",
        "Service Provider", "Provider GSTIN", 
        "Service Receiver", "Receiver GSTIN", 
        "Place of Supply", "Category Code/HSN", "Service Description",
        "Fee Amount (INR)", "CGST (INR)", "SGST (INR)", "IGST (INR)",
        "Total Tax (INR)", "Total Amount (INR)", "Tax Rate %",
        "Template ID", "Validation Status", "Source File"
    ]
    
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        
        for file_data in files_data:
            if file_data.get('status') != 'done':
                continue
            
            data = file_data.get('extraction_data', {})
            line_items = data.get('line_items', [])
            validation = data.get('_validation', {})
            
            # Common fields for all rows
            common = [
                data.get('invoice_number', ''),
                data.get('invoice_date', ''),
                data.get('document_type', ''),
                data.get('source_platform', ''),
                data.get('service_provider_name', ''),
                data.get('service_provider_gstin', ''),
                data.get('service_receiver_name', ''),
                data.get('service_receiver_gstin', ''),
                data.get('place_of_supply_state_code', ''),
            ]
            
            meta = [
                data.get('_template_id', ''),
                validation.get('status', 'ok'),
                data.get('_source_file', file_data.get('original_filename', ''))
            ]
            
            if not line_items:
                # Write header-level totals if no line items
                row = common + [
                    '', '',  # category, description
                    data.get('subtotal_fee_amount', ''),
                    data.get('cgst_amount', ''),
                    data.get('sgst_amount', ''),
                    data.get('igst_amount', ''),
                    data.get('total_tax_amount', ''),
                    data.get('total_invoice_amount', ''),
                    '',  # tax rate
                ] + meta
                writer.writerow(row)
            else:
                for item in line_items:
                    if isinstance(item, dict):
                        row = common + [
                            item.get('category_code_or_hsn', ''),
                            item.get('service_description', ''),
                            item.get('fee_amount', ''),
                            item.get('cgst_amount', ''),
                            item.get('sgst_amount', ''),
                            item.get('igst_amount', ''),
                            item.get('total_tax_amount', ''),
                            item.get('total_amount', ''),
                            item.get('tax_rate_percent', ''),
                        ] + meta
                        writer.writerow(row)
    
    return str(csv_path)


def generate_excel(job_id: str, files_data: List[Dict]) -> str:
    """
    Generate Excel file with multiple sheets:
    - Invoice_Details: Header-level data
    - Line_Items: All line items (long format)
    - Errors_Logs: Validation issues and errors
    """
    excel_path = EXPORT_DIR / f"{job_id}_invoice_data.xlsx"
    
    wb = openpyxl.Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    warn_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    error_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ========== Sheet 1: Invoice_Details ==========
    ws1 = wb.active
    ws1.title = "Invoice_Details"
    
    detail_headers = [
        "Invoice Number", "Invoice Date", "Document Type", "Platform", "Template ID",
        "Provider Name", "Provider GSTIN", "Receiver Name", "Receiver GSTIN",
        "Place of Supply", "Subtotal Fee", "CGST", "SGST", "IGST", 
        "Total Tax", "Total Amount", "Validation Status", "Source File"
    ]
    
    for col, header in enumerate(detail_headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    row_num = 2
    for file_data in files_data:
        if file_data.get('status') != 'done':
            continue
        
        data = file_data.get('extraction_data', {})
        validation = data.get('_validation', {})
        
        values = [
            data.get('invoice_number', ''),
            data.get('invoice_date', ''),
            data.get('document_type', ''),
            data.get('source_platform', ''),
            data.get('_template_id', ''),
            data.get('service_provider_name', ''),
            data.get('service_provider_gstin', ''),
            data.get('service_receiver_name', ''),
            data.get('service_receiver_gstin', ''),
            data.get('place_of_supply_state_code', ''),
            data.get('subtotal_fee_amount') or 0,
            data.get('cgst_amount') or 0,
            data.get('sgst_amount') or 0,
            data.get('igst_amount') or 0,
            data.get('total_tax_amount') or 0,
            data.get('total_invoice_amount') or 0,
            validation.get('status', 'ok'),
            data.get('_source_file', file_data.get('original_filename', ''))
        ]
        
        for col, value in enumerate(values, 1):
            cell = ws1.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            
            # Highlight validation issues
            if col == 17:  # Validation Status column
                if value == 'warn':
                    cell.fill = warn_fill
                elif value == 'fail':
                    cell.fill = error_fill
        
        row_num += 1
    
    # ========== Sheet 2: Line_Items ==========
    ws2 = wb.create_sheet("Line_Items")
    
    item_headers = [
        "Invoice Number", "Invoice Date", "Platform", "Category Code/HSN",
        "Service Description", "Fee Amount", "CGST", "SGST", "IGST",
        "Total Tax", "Total Amount", "Tax Rate %", "Source File"
    ]
    
    for col, header in enumerate(item_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    row_num = 2
    for file_data in files_data:
        if file_data.get('status') != 'done':
            continue
        
        data = file_data.get('extraction_data', {})
        line_items = data.get('line_items', [])
        
        if not line_items:
            # Add header-level as single row
            values = [
                data.get('invoice_number', ''),
                data.get('invoice_date', ''),
                data.get('source_platform', ''),
                '', '',  # category, description
                data.get('subtotal_fee_amount') or 0,
                data.get('cgst_amount') or 0,
                data.get('sgst_amount') or 0,
                data.get('igst_amount') or 0,
                data.get('total_tax_amount') or 0,
                data.get('total_invoice_amount') or 0,
                '',  # tax rate
                data.get('_source_file', '')
            ]
            
            for col, value in enumerate(values, 1):
                cell = ws2.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
            row_num += 1
        else:
            for item in line_items:
                if isinstance(item, dict):
                    values = [
                        data.get('invoice_number', ''),
                        data.get('invoice_date', ''),
                        data.get('source_platform', ''),
                        item.get('category_code_or_hsn', ''),
                        item.get('service_description', ''),
                        item.get('fee_amount') or 0,
                        item.get('cgst_amount') or 0,
                        item.get('sgst_amount') or 0,
                        item.get('igst_amount') or 0,
                        item.get('total_tax_amount') or 0,
                        item.get('total_amount') or 0,
                        item.get('tax_rate_percent') or '',
                        data.get('_source_file', '')
                    ]
                    
                    for col, value in enumerate(values, 1):
                        cell = ws2.cell(row=row_num, column=col, value=value)
                        cell.border = thin_border
                    row_num += 1
    
    # ========== Sheet 3: Errors_Logs ==========
    ws3 = wb.create_sheet("Errors_Logs")
    
    error_headers = ["Source File", "Template ID", "Status", "Issues", "Extraction Method"]
    
    for col, header in enumerate(error_headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    row_num = 2
    for file_data in files_data:
        data = file_data.get('extraction_data', {})
        validation = data.get('_validation', {})
        
        # Only log files with issues or failures
        status = file_data.get('status', '')
        val_status = validation.get('status', 'ok')
        
        if status == 'failed' or val_status in ['warn', 'fail']:
            issues = validation.get('issues', [])
            if status == 'failed':
                issues = [file_data.get('error_message', 'Unknown error')]
            
            values = [
                data.get('_source_file', file_data.get('original_filename', '')),
                data.get('_template_id', 'N/A'),
                'FAILED' if status == 'failed' else val_status.upper(),
                '; '.join(issues) if issues else 'No specific issues',
                data.get('_extraction_method', 'N/A')
            ]
            
            for col, value in enumerate(values, 1):
                cell = ws3.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
                
                if col == 3:  # Status column
                    if 'FAIL' in str(value):
                        cell.fill = error_fill
                    elif 'WARN' in str(value):
                        cell.fill = warn_fill
            
            row_num += 1
    
    # Auto-adjust column widths
    for ws in [ws1, ws2, ws3]:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    wb.save(excel_path)
    return str(excel_path)


# ============== API ROUTES ==============

@invoice_router.get("/")
async def invoice_root():
    return {"message": "Invoice Extractor API - Hybrid Pipeline v2", "status": "ready"}


@invoice_router.post("/upload", response_model=UploadResponse)
async def upload_invoice_files(files: List[UploadFile] = File(...)):
    """Upload multiple PDF files for extraction."""
    if db is None:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    if not files:
        raise HTTPException(status_code=400, detail="No files uploaded")
    
    job_id = str(uuid.uuid4())
    uploaded_files = []
    
    for file in files:
        if not file.filename.lower().endswith('.pdf'):
            raise HTTPException(
                status_code=400, 
                detail=f"Invalid file type: {file.filename}. Only PDF files are allowed."
            )
        
        content = await file.read()
        
        if len(content) > MAX_FILE_SIZE:
            raise HTTPException(
                status_code=400,
                detail=f"File {file.filename} exceeds maximum size of 25MB"
            )
        
        file_id = str(uuid.uuid4())
        safe_filename = f"{file_id}.pdf"
        file_path = UPLOAD_DIR / safe_filename
        
        async with aiofiles.open(file_path, 'wb') as f:
            await f.write(content)
        
        file_info = {
            "id": file_id,
            "job_id": job_id,
            "filename": safe_filename,
            "original_filename": file.filename,
            "status": "pending",
            "error_message": None,
            "extraction_data": None,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db.invoice_job_files.insert_one(file_info)
        
        uploaded_files.append({
            "id": file_id,
            "filename": file.filename,
            "status": "pending"
        })
    
    job_record = {
        "job_id": job_id,
        "total_files": len(uploaded_files),
        "processed_files": 0,
        "failed_files": 0,
        "status": "idle",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.invoice_jobs.insert_one(job_record)
    
    return UploadResponse(job_id=job_id, files=uploaded_files)


@invoice_router.post("/extract/{job_id}")
async def start_invoice_extraction(job_id: str, background_tasks: BackgroundTasks, use_llm: bool = True):
    """Start the extraction process for a job."""
    if db is None:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    job = await db.invoice_jobs.find_one({"job_id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    if job.get('status') == 'processing':
        raise HTTPException(status_code=400, detail="Job is already processing")
    
    api_key = os.environ.get('EMERGENT_LLM_KEY')
    background_tasks.add_task(run_extraction_job, job_id, use_llm, api_key)
    
    return {"message": "Extraction started", "job_id": job_id, "use_llm": use_llm}


@invoice_router.get("/job/{job_id}", response_model=JobStatusResponse)
async def get_invoice_job_status(job_id: str):
    """Get the status of an extraction job."""
    if db is None:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    job = await db.invoice_jobs.find_one({"job_id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    files = await db.invoice_job_files.find({"job_id": job_id}, {"_id": 0}).to_list(1000)
    
    return JobStatusResponse(
        job_id=job_id,
        status=job.get('status', 'idle'),
        total_files=job.get('total_files', 0),
        processed_files=job.get('processed_files', 0),
        failed_files=job.get('failed_files', 0),
        files=[{
            "id": f.get('id'),
            "filename": f.get('original_filename'),
            "status": f.get('status'),
            "error_message": f.get('error_message'),
            "extraction_data": f.get('extraction_data')
        } for f in files]
    )


@invoice_router.get("/export/csv/{job_id}")
async def export_invoice_csv(job_id: str):
    """Export extraction results as CSV."""
    if db is None:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    job = await db.invoice_jobs.find_one({"job_id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    if job.get('status') != 'completed':
        raise HTTPException(status_code=400, detail="Job not completed yet")
    
    files = await db.invoice_job_files.find({"job_id": job_id}, {"_id": 0}).to_list(1000)
    csv_path = generate_csv(job_id, files)
    
    def iterfile():
        with open(csv_path, 'rb') as f:
            yield from f
    
    filename = f"invoice_data_{job_id[:8]}.csv"
    return StreamingResponse(
        iterfile(),
        media_type="text/csv",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )


@invoice_router.get("/export/excel/{job_id}")
async def export_invoice_excel(job_id: str):
    """Export extraction results as Excel."""
    if db is None:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    job = await db.invoice_jobs.find_one({"job_id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    if job.get('status') != 'completed':
        raise HTTPException(status_code=400, detail="Job not completed yet")
    
    files = await db.invoice_job_files.find({"job_id": job_id}, {"_id": 0}).to_list(1000)
    excel_path = generate_excel(job_id, files)
    
    def iterfile():
        with open(excel_path, 'rb') as f:
            yield from f
    
    filename = f"invoice_data_{job_id[:8]}.xlsx"
    return StreamingResponse(
        iterfile(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )


@invoice_router.delete("/job/{job_id}")
async def delete_invoice_job(job_id: str):
    """Delete a job and its associated files."""
    if db is None:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    job = await db.invoice_jobs.find_one({"job_id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    files = await db.invoice_job_files.find({"job_id": job_id}, {"_id": 0}).to_list(1000)
    
    for file_info in files:
        file_path = UPLOAD_DIR / file_info.get('filename', '')
        if file_path.exists():
            file_path.unlink()
    
    csv_path = EXPORT_DIR / f"{job_id}_invoice_line_items.csv"
    excel_path = EXPORT_DIR / f"{job_id}_invoice_data.xlsx"
    if csv_path.exists():
        csv_path.unlink()
    if excel_path.exists():
        excel_path.unlink()
    
    await db.invoice_job_files.delete_many({"job_id": job_id})
    await db.invoice_jobs.delete_one({"job_id": job_id})
    
    return {"message": "Job deleted successfully"}
